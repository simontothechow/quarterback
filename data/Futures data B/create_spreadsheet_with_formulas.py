"""
Script to recreate the Futures Mid Spreads spreadsheet WITH FORMULAS.
Dates are stored as actual Excel date values (not text) so formulas work correctly.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mid Spreads"

# Define colors
RED_FILL = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
LIGHT_ORANGE_FILL = PatternFill(start_color="FFCC80", end_color="FFCC80", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")
DARK_FILL = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF")

# Row 2: Days to expiry for each "to" contract column
# These values represent days from a reference point to each contract's expiry
row2_data = {
    'E': -42, 'F': -7, 'G': 21, 'H': 49, 'I': 77, 'J': 111,
    'K': 203, 'L': 294, 'M': 385, 'N': 475, 'O': 567, 'P': 658
}
for col, val in row2_data.items():
    ws[f'{col}2'] = val

# Row 3: Contract codes for each column (the "to" contracts)
ws['D3'] = "End"
row3_contracts = ['AXWF6', 'AXWG6', 'AXWH6', 'AXWJ6', 'AXWK6', 'AXWM6', 
                  'AXWU6', 'AXWZ6', 'AXWH7', 'AXWM7', 'AXWU7', 'AXWZ7']
for i, code in enumerate(row3_contracts):
    col = get_column_letter(5 + i)  # Starting from E
    ws[f'{col}3'] = code

# Row 4: Spread metric values for each "to" contract
row4_data = {
    'E': 116, 'F': 'VALUEERROR', 'G': 44.5, 'H': 48, 'I': 110, 'J': 99,
    'K': 77, 'L': 55.5, 'M': 81.5, 'N': 66.5, 'O': 69, 'P': 75
}
for col, val in row4_data.items():
    ws[f'{col}4'] = val

# Row 5: Headers (A-D) and DELIVERY DATES as actual Excel dates (E-P)
ws['A5'] = "Start"
# B5 is empty (contract code column)
ws['C5'] = "last_price"
ws['D5'] = "FUT_DLV_DT_FIRST"

# Delivery dates as actual datetime objects (Excel will recognize these as dates)
row5_dates = [
    datetime(2026, 1, 16),   # E5 - AXWF6
    datetime(2026, 2, 20),   # F5 - AXWG6
    datetime(2026, 3, 20),   # G5 - AXWH6
    datetime(2026, 4, 17),   # H5 - AXWJ6
    datetime(2026, 5, 15),   # I5 - AXWK6
    datetime(2026, 6, 18),   # J5 - AXWM6
    datetime(2026, 9, 18),   # K5 - AXWU6
    datetime(2026, 12, 18),  # L5 - AXWZ6
    datetime(2027, 3, 19),   # M5 - AXWH7
    datetime(2027, 6, 17),   # N5 - AXWM7
    datetime(2027, 9, 17),   # O5 - AXWU7
    datetime(2027, 12, 17),  # P5 - AXWZ7
]
for i, date_val in enumerate(row5_dates):
    col = get_column_letter(5 + i)  # Starting from E
    cell = ws[f'{col}5']
    cell.value = date_val
    cell.number_format = 'M/D/YYYY'

# Data rows: Each row represents a "from" contract
# Column A: Start (days to expiry), Column B: Contract code, 
# Column C: last_price, Column D: FUT_DLV_DT_FIRST (as actual date)

row_data = [
    # (Start, Contract, last_price, delivery_date)
    (-70, 'AXWZ5', 65, datetime(2025, 12, 19)),      # Row 6
    (-42, 'AXWF6', 'VALUEERROR', datetime(2026, 1, 16)),  # Row 7 - #N/A becomes VALUEERROR
    (-7, 'AXWG6', 'VALUEERROR', datetime(2026, 2, 20)),   # Row 8
    (21, 'AXWH6', 44.5, datetime(2026, 3, 20)),      # Row 9
    (49, 'AXWJ6', 72, datetime(2026, 4, 17)),        # Row 10
    (77, 'AXWK6', 95, datetime(2026, 5, 15)),        # Row 11
    (111, 'AXWM6', 51.5, datetime(2026, 6, 18)),     # Row 12
    (203, 'AXWU6', 54.5, datetime(2026, 9, 18)),     # Row 13
    (294, 'AXWZ6', 56, datetime(2026, 12, 18)),      # Row 14
    (385, 'AXWH7', 77.5, datetime(2027, 3, 19)),     # Row 15
    (475, 'AXWM7', 66.5, datetime(2027, 6, 17)),     # Row 16
    (567, 'AXWU7', 69, datetime(2027, 9, 17)),       # Row 17
    (658, 'AXWZ7', 75, datetime(2027, 12, 17)),      # Row 18
    (749, 'AXWH8', 70, datetime(2028, 3, 17)),       # Row 19
    (840, 'AXWM8', 74.5, datetime(2028, 6, 16)),     # Row 20
    (931, 'AXWU8', 88, datetime(2028, 9, 15)),       # Row 21
    (1022, 'AXWZ8', 74, datetime(2028, 12, 15)),     # Row 22
    (1393, 'AXWH9', 80, datetime(2029, 12, 21)),     # Row 23
    (1757, 'AXWZ0', 88, datetime(2030, 12, 20)),     # Row 24 - Yellow highlight
    (2121, 'AXWZ31', 94.5, datetime(2031, 12, 19)),  # Row 25
    (2485, 'AXWZ32', 103.5, datetime(2032, 12, 17)), # Row 26
    (2849, 'AXWZ33', 115, datetime(2033, 12, 16)),   # Row 27
    (3213, 'AXWZ34', 120, datetime(2034, 12, 15)),   # Row 28
    (3584, 'AXWZ35', 122, datetime(2035, 12, 21)),   # Row 29
]

for i, (start, contract, price, delivery_date) in enumerate(row_data):
    row = 6 + i
    ws[f'A{row}'] = start
    ws[f'B{row}'] = contract
    ws[f'C{row}'] = price
    
    # Delivery date as actual date
    date_cell = ws[f'D{row}']
    date_cell.value = delivery_date
    date_cell.number_format = 'M/D/YYYY'
    
    # Highlight row 24 (AXWZ0) in yellow
    if row == 24:
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].fill = YELLOW_FILL

# Now create the FORMULAS for the spread grid (E6 onwards)
# Formula pattern: =([col]4-($C$[row]*(($D$[row]-(TODAY()))/($[col]5-(TODAY())))))/(($[col]2-$A$[row])/[col]2)
#
# The formula calculates:
# (To_Spread_Value - (From_Price × (From_Days_Remaining / To_Days_Remaining))) / ((To_DTE - From_DTE) / To_DTE)

# Map of which cells should have formulas vs be dark (empty)
# A cell should be dark if the "from" contract expires AFTER or SAME as the "to" contract
# i.e., when $A$[row] >= [col]2

# Column letters for E through P
spread_columns = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']

# Days to expiry for each column (from row 2)
col_dte = {'E': -42, 'F': -7, 'G': 21, 'H': 49, 'I': 77, 'J': 111, 
           'K': 203, 'L': 294, 'M': 385, 'N': 475, 'O': 567, 'P': 658}

# Days to expiry for each row (from column A)
row_dte = {6: -70, 7: -42, 8: -7, 9: 21, 10: 49, 11: 77, 12: 111, 
           13: 203, 14: 294, 15: 385, 16: 475, 17: 567, 18: 658,
           19: 749, 20: 840, 21: 931, 22: 1022, 23: 1393, 24: 1757,
           25: 2121, 26: 2485, 27: 2849, 28: 3213, 29: 3584}

# Rows with VALUEERROR in last_price (can't calculate spreads)
valueerror_rows = [7, 8]

for row in range(6, 30):  # Rows 6 to 29
    for col in spread_columns:
        cell = ws[f'{col}{row}']
        
        from_dte = row_dte.get(row, 9999)
        to_dte = col_dte.get(col, -9999)
        
        # Check if this cell should be dark (from contract expires after to contract)
        # or if the from contract has no valid price
        if from_dte >= to_dte or row in valueerror_rows:
            cell.fill = DARK_FILL
        else:
            # Create the formula - MUST match original exactly:
            # =(J4-($C$9*(($D$9-(TODAY()))/($J5-(TODAY())))))/(($J2-$A$9)/J2)
            # Note the parentheses around TODAY() and the denominator structure
            formula = f'=({col}4-($C${row}*(($D${row}-(TODAY()))/({col}$5-(TODAY())))))/(({col}$2-$A${row})/{col}$2)'
            cell.value = formula
            cell.number_format = '0.00000'

# Special handling for certain highlighted cells based on the original screenshots
# Orange highlights on diagonal-ish cells
orange_cells = ['I9', 'N16', 'O17']
for cell_ref in orange_cells:
    ws[cell_ref].fill = ORANGE_FILL

# Light orange for specific cells
light_orange_cells = ['M15']
for cell_ref in light_orange_cells:
    ws[cell_ref].fill = LIGHT_ORANGE_FILL

# Row 34: Steepness label
ws['C34'] = "Steepness"

# Row 35: Steepness formula (calculates slope of the curve)
# This appears to show -13.40% in the original - likely a calculation of term structure steepness
# For now, put a placeholder formula or the observed value
ws['O35'] = -0.1340
ws['O35'].number_format = '0.00%'
ws['O35'].fill = ORANGE_FILL

# Adjust column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 14
for col in spread_columns:
    ws.column_dimensions[col].width = 12

# Save workbook
output_path = r"C:\Users\simon\OneDrive\Desktop\Quarterback_v1\data\Futures data B\Mid_Spreads_With_Formulas.xlsx"
wb.save(output_path)
print(f"Spreadsheet saved to: {output_path}")
print("\nKey features:")
print("- Dates stored as actual Excel date values (not text)")
print("- Formulas in spread grid cells (E6:P29)")
print("- Dark cells where from_contract >= to_contract (no valid spread)")
print("- VALUEERROR rows (7, 8) have dark cells (no price data)")
print("\nFormula used:")
print("=([col]4-($C$[row]*(($D$[row]-TODAY())/([col]$5-TODAY()))))/((${col}$2-$A$[row])/[col]$2)")
