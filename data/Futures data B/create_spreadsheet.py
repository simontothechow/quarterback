"""
Script to recreate the Futures Mid Spreads spreadsheet from screenshots.
Values marked with UNCERTAIN comment will be highlighted in purple.
"""

import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mid Spreads"

# Define colors
RED_FILL = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
LIGHT_ORANGE_FILL = PatternFill(start_color="FFCC80", end_color="FFCC80", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")  # Uncertain values
DARK_FILL = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
LIGHT_YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

# Row 2: Column numbers (days to expiry differences)
row2_data = {
    'E': -42, 'F': -7, 'G': 21, 'H': 49, 'I': 77, 'J': 111,
    'K': 203, 'L': 294, 'M': 385, 'N': 475, 'O': 567, 'P': 658, 'Q': 749, 'R': 840
}
for col, val in row2_data.items():
    ws[f'{col}2'] = val

# Row 3: Contract codes
ws['D3'] = "End"
row3_contracts = ['AXWF6', 'AXWG6', 'AXWH6', 'AXWJ6', 'AXWK6', 'AXWM6', 'AXWU6', 'AXWZ6', 
                  'AXWH7', 'AXWM7', 'AXWU7', 'AXWZ7', 'AXWH8', 'AXWM8']
for i, code in enumerate(row3_contracts):
    col = get_column_letter(5 + i)  # Starting from E
    ws[f'{col}3'] = code

# Row 4: Additional values (days or other metrics)
row4_data = {
    'E': 116, 'F': '#N/A Real T', 'G': 44.5, 'H': 48, 'I': 110, 'J': 99,
    'K': 77, 'L': 55.5, 'M': 81.5, 'N': 66.5, 'O': 65.5, 'P': 74, 'Q': 65.5, 'R': 74
}
# Note: F4 is uncertain, P4, Q4, R4 are uncertain
uncertain_row4 = ['F', 'P', 'Q', 'R']
for col, val in row4_data.items():
    ws[f'{col}4'] = val
    if col in uncertain_row4:
        ws[f'{col}4'].fill = PURPLE_FILL

# Row 5: Headers and dates
# A5=Start, B5=(empty - contract code column), C5=last_price, D5=FUT_DLV_DT_FIRST
# E5 onwards = dates for each contract column
ws['A5'] = "Start"
# B5 is empty (contract code column doesn't need a header in row 5)
ws['C5'] = "last_price"
ws['D5'] = "FUT_DLV_DT_FIRST"

# Dates in columns E through P (E5 to P5)
row5_dates = ['1/16/2026', '2/20/2026', '3/20/2026', '4/17/2026', '5/15/2026', '6/18/2026',
              '9/18/2026', '12/18/2026', '3/19/2027', '6/17/2027', '9/17/2027', '12/17/2027']
for i, date in enumerate(row5_dates):
    col = get_column_letter(5 + i)  # Starting from E (5th column)
    ws[f'{col}5'] = date

# Data rows (6-29)
# Column A: Start (days), Column B: Contract, Column C: last_price, Column D: FUT_DLV_DT_FIRST
row_data = [
    # Row 6: AXWZ5
    {'A': -70, 'B': 'AXWZ5', 'C': 65, 'D': '12/19/2025'},
    # Row 7: AXWF6
    {'A': -42, 'B': 'AXWF6', 'C': '#N/A', 'D': 'Real Tir 1/16/2026'},
    # Row 8: AXWG6
    {'A': -7, 'B': 'AXWG6', 'C': '#N/A', 'D': 'Real Tir 2/20/2026'},
    # Row 9: AXWH6
    {'A': 21, 'B': 'AXWH6', 'C': 44.5, 'D': '3/20/2026'},
    # Row 10: AXWJ6
    {'A': 49, 'B': 'AXWJ6', 'C': 72, 'D': '4/17/2026'},
    # Row 11: AXWK6
    {'A': 77, 'B': 'AXWK6', 'C': 95, 'D': '5/15/2026'},
    # Row 12: AXWM6
    {'A': 111, 'B': 'AXWM6', 'C': 51.5, 'D': '6/18/2026'},
    # Row 13: AXWU6
    {'A': 203, 'B': 'AXWU6', 'C': 54.5, 'D': '9/18/2026'},
    # Row 14: AXWZ6
    {'A': 294, 'B': 'AXWZ6', 'C': 56, 'D': '12/18/2026'},
    # Row 15: AXWH7
    {'A': 385, 'B': 'AXWH7', 'C': 77.5, 'D': '3/19/2027'},
    # Row 16: AXWM7
    {'A': 475, 'B': 'AXWM7', 'C': 66.5, 'D': '6/17/2027'},
    # Row 17: AXWU7
    {'A': 567, 'B': 'AXWU7', 'C': 69, 'D': '9/17/2027'},
    # Row 18: AXWZ7
    {'A': 658, 'B': 'AXWZ7', 'C': 75, 'D': '12/17/2027'},
    # Row 19: AXWH8
    {'A': 749, 'B': 'AXWH8', 'C': 70, 'D': '3/17/2028'},
    # Row 20: AXWM8
    {'A': 840, 'B': 'AXWM8', 'C': 74.5, 'D': '6/16/2028'},
    # Row 21: AXWU8
    {'A': 931, 'B': 'AXWU8', 'C': 88, 'D': '9/15/2028'},
    # Row 22: AXWZ8
    {'A': 1022, 'B': 'AXWZ8', 'C': 74, 'D': '12/15/2028'},
    # Row 23: AXWH9
    {'A': 1393, 'B': 'AXWH9', 'C': 80, 'D': '12/21/2029'},
    # Row 24: AXWZ0 (highlighted row - 10 year point)
    {'A': 1757, 'B': 'AXWZ0', 'C': 88, 'D': '12/20/2030'},
    # Row 25: AXWZ31
    {'A': 2121, 'B': 'AXWZ31', 'C': 94.5, 'D': '12/19/2031'},
    # Row 26: AXWZ32
    {'A': 2485, 'B': 'AXWZ32', 'C': 103.5, 'D': '12/17/2032'},
    # Row 27: AXWZ33
    {'A': 2849, 'B': 'AXWZ33', 'C': 115, 'D': '12/16/2033'},
    # Row 28: AXWZ34
    {'A': 3213, 'B': 'AXWZ34', 'C': 120, 'D': '12/15/2034'},
    # Row 29: AXWZ35
    {'A': 3584, 'B': 'AXWZ35', 'C': 122, 'D': '12/21/2035'},
]

for i, data in enumerate(row_data):
    row = 6 + i
    for col, val in data.items():
        ws[f'{col}{row}'] = val
    # Highlight row 24 (AXWZ0) in yellow
    if row == 24:
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].fill = YELLOW_FILL

# Now the spread values grid (E6 onwards)
# This is the complex part - extracting values from the images

# Row 6 (AXWZ5) spread values
row6_spreads = {
    'E': (-11.50000, 'red'),
    'F': ('VALUEERROR', 'red'),
    'G': (60.26923, None),
    'H': (58.00000, None),
    'I': (88.57143, None),
    'J': (85.85083, None),
    'K': (73.92308, None),
    'L': (57.32692, None),
    'M': (78.96154, None),
    'N': (66.30734, None),
    'O': (68.56044, None),
    'P': (65.45192, None),
    'Q': (83.29060, 'uncertain'),
    'R': (73.7692, 'uncertain'),
}

# Row 7 (AXWF6) - all VALUEERROR
row7_spreads = {col: ('VALUEERROR', 'red') for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']}

# Row 8 (AXWG6) - all VALUEERROR
row8_spreads = {col: ('VALUEERROR', 'red') for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']}

# Row 9 (AXWH6)
row9_spreads = {
    'E': (None, 'dark'),
    'F': (None, 'dark'),
    'G': (50.62500, None),
    'H': (134.56250, None),
    'I': (111.71667, 'orange'),  # Highlighted cell - formula result
    'J': (80.75000, None),
    'K': (56.34615, None),
    'L': (83.63462, None),
    'M': (67.51762, None),
    'N': (69.94231, None),
    'O': (66.19231, None),
    'P': (86.16827, 'uncertain'),
    'Q': (75.2692, 'uncertain'),
    'R': (None, 'uncertain'),
}

# Row 10 (AXWJ6)
row10_spreads = {
    'E': (None, 'dark'),
    'F': (None, 'dark'),
    'G': (None, 'dark'),
    'H': (176.50000, None),
    'I': (120.33871, None),
    'J': (78.59091, None),
    'K': (52.20000, None),
    'L': (82.88542, None),
    'M': (65.86737, None),
    'N': (68.71622, None),
    'O': (64.97701, None),
    'P': (85.91000, None),
    'Q': (74.6543, 'uncertain'),
    'R': (None, 'uncertain'),
}

# Row 11 (AXWK6)
row11_spreads = {
    'E': (None, 'dark'),
    'F': (None, 'dark'),
    'G': (None, 'dark'),
    'H': (None, 'dark'),
    'I': (108.05882, 'uncertain'),
    'J': (66.00000, None),
    'K': (41.48387, None),
    'L': (78.12500, None),
    'M': (60.98618, None),
    'N': (64.91429, None),
    'O': (61.59036, 'uncertain'),
    'P': (83.85417, 'uncertain'),
    'Q': (72.43117, 'uncertain'),
    'R': (None, 'uncertain'),
}

# Row 12 (AXWM6)
row12_spreads = {
    'E': (None, 'dark'),
    'F': (None, 'dark'),
    'G': (None, 'dark'),
    'H': (None, 'dark'),
    'I': (None, 'dark'),
    'J': (107.76630, 'uncertain'),
    'K': (57.73077, 'uncertain'),
    'L': (93.65328, None),
    'M': (71.07418, None),
    'N': (75.25987, 'uncertain'),
    'O': (68.34095, None),
    'P': (90.82877, 'uncertain'),
    'Q': (78.02317, 'uncertain'),
    'R': (None, 'uncertain'),
}

# Row 13 (AXWU6)
row13_spreads = {
    'E': (None, 'dark'),
    'F': (None, 'dark'),
    'G': (None, 'dark'),
    'H': (None, 'dark'),
    'I': (None, 'dark'),
    'J': (None, 'dark'),
    'K': (57.73077, 'uncertain'),
    'L': (111.61538, None),
    'M': (75.45588, None),
    'N': (77.08654, 'uncertain'),
    'O': (70.40769, None),
    'P': (96.33974, 'uncertain'),
    'Q': (80.8730, 'uncertain'),
    'R': (None, 'uncertain'),
}

# Row 14 (AXWZ6)
row14_spreads = {
    'E': (None, 'dark'),
    'F': (None, 'dark'),
    'G': (None, 'dark'),
    'H': (None, 'dark'),
    'I': (None, 'dark'),
    'J': (None, 'dark'),
    'K': (None, 'dark'),
    'L': (163.88462, 'uncertain'),
    'M': (83.55525, None),
    'N': (83.00000, 'uncertain'),
    'O': (73.17308, 'uncertain'),
    'P': (103.73846, 'uncertain'),
    'Q': (84.4615, 'uncertain'),
    'R': (None, 'uncertain'),
}

# Row 15 (AXWH7)
row15_spreads = {
    'E': (None, 'dark'),
    'F': (None, 'dark'),
    'G': (None, 'dark'),
    'H': (None, 'dark'),
    'I': (None, 'dark'),
    'J': (None, 'dark'),
    'K': (None, 'dark'),
    'L': (None, 'dark'),
    'M': (19.44444, 'light_orange'),
    'N': (51.01923, None),
    'O': (48.57692, 'uncertain'),
    'P': (92.93269, 'uncertain'),
    'Q': (71.9615, 'uncertain'),
    'R': (None, 'uncertain'),
}

# Row 16 (AXWM7)
row16_spreads = {
    'E': (None, 'dark'),
    'F': (None, 'dark'),
    'G': (None, 'dark'),
    'H': (None, 'dark'),
    'I': (None, 'dark'),
    'J': (None, 'dark'),
    'K': (None, 'dark'),
    'L': (None, 'dark'),
    'M': (None, 'dark'),
    'N': (81.90761, 'orange'),
    'O': (62.90437, None),
    'P': (117.07117, 'uncertain'),
    'Q': (84.9108, 'uncertain'),
    'R': (None, 'uncertain'),
}

# Row 17 (AXWU7)
row17_spreads = {
    'E': (None, 'dark'),
    'F': (None, 'dark'),
    'G': (None, 'dark'),
    'H': (None, 'dark'),
    'I': (None, 'dark'),
    'J': (None, 'dark'),
    'K': (None, 'dark'),
    'L': (None, 'dark'),
    'M': (None, 'dark'),
    'N': (None, 'dark'),
    'O': (43.69231, 'orange'),
    'P': (134.84615, 'uncertain'),
    'Q': (85.9270, 'uncertain'),
    'R': (None, 'uncertain'),
}

# Row 18 (AXWZ7)
row18_spreads = {
    'E': (None, 'dark'),
    'F': (None, 'dark'),
    'G': (None, 'dark'),
    'H': (None, 'dark'),
    'I': (None, 'dark'),
    'J': (None, 'dark'),
    'K': (None, 'dark'),
    'L': (None, 'dark'),
    'M': (None, 'dark'),
    'N': (None, 'dark'),
    'O': (None, 'dark'),
    'P': (157.30769, 'uncertain'),
    'Q': (72.8923, 'uncertain'),
    'R': (None, 'uncertain'),
}

# Row 19 (AXWH8)
row19_spreads = {
    'E': (None, 'dark'),
    'F': (None, 'dark'),
    'G': (None, 'dark'),
    'H': (None, 'dark'),
    'I': (None, 'dark'),
    'J': (None, 'dark'),
    'K': (None, 'dark'),
    'L': (None, 'dark'),
    'M': (None, 'dark'),
    'N': (None, 'dark'),
    'O': (None, 'dark'),
    'P': (None, 'dark'),
    'Q': (111.53846, 'uncertain'),
    'R': (None, 'uncertain'),
}

# Rows 20-29: Mostly dark cells with very few values visible
row20_29_spreads = {}  # Will be filled with dark cells mostly

# Compile all spread data
all_spreads = {
    6: row6_spreads,
    7: row7_spreads,
    8: row8_spreads,
    9: row9_spreads,
    10: row10_spreads,
    11: row11_spreads,
    12: row12_spreads,
    13: row13_spreads,
    14: row14_spreads,
    15: row15_spreads,
    16: row16_spreads,
    17: row17_spreads,
    18: row18_spreads,
    19: row19_spreads,
}

# Add dark cells for rows 20-29
for row in range(20, 30):
    all_spreads[row] = {col: (None, 'dark') for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']}

# Write spread data to worksheet
for row, spreads in all_spreads.items():
    for col, (val, style) in spreads.items():
        cell = ws[f'{col}{row}']
        if val is not None:
            cell.value = val
        
        # Apply styling
        if style == 'red':
            cell.fill = RED_FILL
        elif style == 'orange':
            cell.fill = ORANGE_FILL
        elif style == 'light_orange':
            cell.fill = LIGHT_ORANGE_FILL
        elif style == 'dark':
            cell.fill = DARK_FILL
        elif style == 'uncertain':
            cell.fill = PURPLE_FILL
        elif style == 'yellow':
            cell.fill = YELLOW_FILL

# Row 34: Steepness label
ws['C34'] = "Steepness"

# Row 35: Steepness value (seen in lower right of image)
ws['O35'] = "-13.40%"
ws['O35'].fill = ORANGE_FILL

# Adjust column widths
for col in range(1, 19):
    ws.column_dimensions[get_column_letter(col)].width = 12

# Save workbook
output_path = r"C:\Users\simon\OneDrive\Desktop\Quarterback_v1\data\Futures data B\Mid_Spreads_Reconstructed.xlsx"
wb.save(output_path)
print(f"Spreadsheet saved to: {output_path}")
print("\nLegend:")
print("- PURPLE cells: Values I was uncertain about (hard to read)")
print("- RED cells: VALUEERROR or negative values")
print("- ORANGE cells: Highlighted values from original")
print("- DARK GRAY cells: Empty/no value cells")
print("- YELLOW cells: Highlighted row (AXWZ0)")
