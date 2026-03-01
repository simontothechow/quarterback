"""
Script to create CORRECTED Futures Mid Spreads spreadsheet.
Key fix: Row 4 now matches Column C values (same contract = same price)
This correctly calculates IMPLIED FORWARD RATES between contract periods.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mid Spreads"

# Define colors
RED_FILL = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
LIGHT_ORANGE_FILL = PatternFill(start_color="FFCC80", end_color="FFCC80", fill_type="solid")
DARK_FILL = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

# Contract data - single source of truth for prices
# Each contract has: (days_to_expiry, price, delivery_date)
contracts = {
    'AXWZ5':  (-70, 65, datetime(2025, 12, 19)),
    'AXWF6':  (-42, 'VALUEERROR', datetime(2026, 1, 16)),
    'AXWG6':  (-7, 'VALUEERROR', datetime(2026, 2, 20)),
    'AXWH6':  (21, 44.5, datetime(2026, 3, 20)),
    'AXWJ6':  (49, 72, datetime(2026, 4, 17)),
    'AXWK6':  (77, 95, datetime(2026, 5, 15)),
    'AXWM6':  (111, 51.5, datetime(2026, 6, 18)),
    'AXWU6':  (203, 54.5, datetime(2026, 9, 18)),
    'AXWZ6':  (294, 56, datetime(2026, 12, 18)),
    'AXWH7':  (385, 77.5, datetime(2027, 3, 19)),
    'AXWM7':  (475, 66.5, datetime(2027, 6, 17)),
    'AXWU7':  (567, 69, datetime(2027, 9, 17)),
    'AXWZ7':  (658, 75, datetime(2027, 12, 17)),
    'AXWH8':  (749, 70, datetime(2028, 3, 17)),
    'AXWM8':  (840, 74.5, datetime(2028, 6, 16)),
    'AXWU8':  (931, 88, datetime(2028, 9, 15)),
    'AXWZ8':  (1022, 74, datetime(2028, 12, 15)),
    'AXWH9':  (1393, 80, datetime(2029, 12, 21)),
    'AXWZ0':  (1757, 88, datetime(2030, 12, 20)),
    'AXWZ31': (2121, 94.5, datetime(2031, 12, 19)),
    'AXWZ32': (2485, 103.5, datetime(2032, 12, 17)),
    'AXWZ33': (2849, 115, datetime(2033, 12, 16)),
    'AXWZ34': (3213, 120, datetime(2034, 12, 15)),
    'AXWZ35': (3584, 122, datetime(2035, 12, 21)),
}

# Column contracts (the "TO" contracts for forward rate calculation)
column_contracts = ['AXWF6', 'AXWG6', 'AXWH6', 'AXWJ6', 'AXWK6', 'AXWM6',
                    'AXWU6', 'AXWZ6', 'AXWH7', 'AXWM7', 'AXWU7', 'AXWZ7']

# Row contracts (the "FROM" contracts for forward rate calculation)
row_contracts = ['AXWZ5', 'AXWF6', 'AXWG6', 'AXWH6', 'AXWJ6', 'AXWK6', 'AXWM6',
                 'AXWU6', 'AXWZ6', 'AXWH7', 'AXWM7', 'AXWU7', 'AXWZ7',
                 'AXWH8', 'AXWM8', 'AXWU8', 'AXWZ8', 'AXWH9', 'AXWZ0',
                 'AXWZ31', 'AXWZ32', 'AXWZ33', 'AXWZ34', 'AXWZ35']

# === ROW 2: Days to expiry for each column contract ===
for i, contract in enumerate(column_contracts):
    col = get_column_letter(5 + i)  # Starting from E
    ws[f'{col}2'] = contracts[contract][0]  # days_to_expiry

# === ROW 3: Contract codes for columns ===
ws['D3'] = "End"
for i, contract in enumerate(column_contracts):
    col = get_column_letter(5 + i)  # Starting from E
    ws[f'{col}3'] = contract

# === ROW 4: Contract PRICES for columns (CORRECTED - matches Column C) ===
for i, contract in enumerate(column_contracts):
    col = get_column_letter(5 + i)  # Starting from E
    price = contracts[contract][1]  # Get price from single source of truth
    ws[f'{col}4'] = price
    # Highlight to show these are now corrected
    if price != 'VALUEERROR':
        ws[f'{col}4'].fill = GREEN_FILL  # Green to show corrected values

# === ROW 5: Headers and delivery dates ===
ws['A5'] = "Start"
ws['C5'] = "last_price"
ws['D5'] = "FUT_DLV_DT_FIRST"

# Delivery dates for column contracts
for i, contract in enumerate(column_contracts):
    col = get_column_letter(5 + i)  # Starting from E
    cell = ws[f'{col}5']
    cell.value = contracts[contract][2]  # delivery_date
    cell.number_format = 'M/D/YYYY'

# === DATA ROWS (6 onwards): Row contracts ===
for i, contract in enumerate(row_contracts):
    row = 6 + i
    days, price, delivery = contracts[contract]
    
    ws[f'A{row}'] = days
    ws[f'B{row}'] = contract
    ws[f'C{row}'] = price
    
    date_cell = ws[f'D{row}']
    date_cell.value = delivery
    date_cell.number_format = 'M/D/YYYY'
    
    # Highlight AXWZ0 row (10-year point) in yellow
    if contract == 'AXWZ0':
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].fill = YELLOW_FILL

# === SPREAD GRID: Formulas for implied forward rates ===
# Formula: =(col4-($C$row*(($D$row-(TODAY()))/($col5-(TODAY())))))/(($col2-$A$row)/$col2)
# This calculates: Implied forward rate from ROW contract to COLUMN contract

spread_columns = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']

for row_idx, row_contract in enumerate(row_contracts):
    row = 6 + row_idx
    row_days = contracts[row_contract][0]
    row_price = contracts[row_contract][1]
    
    for col_idx, col_contract in enumerate(column_contracts):
        col = spread_columns[col_idx]
        col_days = contracts[col_contract][0]
        col_price = contracts[col_contract][1]
        
        cell = ws[f'{col}{row}']
        
        # Dark cell conditions:
        # 1. Row contract expires after or same as column contract (no valid forward period)
        # 2. Either contract has VALUEERROR price
        if row_days >= col_days or row_price == 'VALUEERROR' or col_price == 'VALUEERROR':
            cell.fill = DARK_FILL
        else:
            # Create the forward rate formula
            formula = f'=({col}4-($C${row}*(($D${row}-(TODAY()))/({col}$5-(TODAY())))))/(({col}$2-$A${row})/{col}$2)'
            cell.value = formula
            cell.number_format = '0.00000'

# === Special highlighting for key cells ===
# These show important forward rate periods
highlight_cells = {
    'I9': ORANGE_FILL,   # Mar→May forward
    'M15': LIGHT_ORANGE_FILL,  # Notable forward
    'N16': ORANGE_FILL,  # Jun→Jun forward (1 year)
    'O17': ORANGE_FILL,  # Sep→Sep forward (1 year)
}
for cell_ref, fill in highlight_cells.items():
    if ws[cell_ref].value:  # Only if cell has a formula
        ws[cell_ref].fill = fill

# === ROW 34-35: Steepness calculation ===
ws['C34'] = "Steepness"
ws['O35'] = -0.1340
ws['O35'].number_format = '0.00%'
ws['O35'].fill = ORANGE_FILL

# === Column widths ===
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 14
for col in spread_columns:
    ws.column_dimensions[col].width = 12

# === Add a note about the correction ===
ws['A1'] = "CORRECTED: Row 4 now matches Column C (same contract = same price)"
ws['A1'].font = Font(bold=True, color="008000")

# Save workbook
output_path = r"C:\Users\simon\OneDrive\Desktop\Quarterback_v1\data\Futures data B\Mid_Spreads_Corrected.xlsx"
wb.save(output_path)

print(f"Spreadsheet saved to: {output_path}")
print("\n" + "="*60)
print("CORRECTIONS MADE:")
print("="*60)
print("\nRow 4 values now match Column C (single source of truth):")
print("-" * 50)
print(f"{'Contract':<10} {'Old Row 4':<12} {'New Row 4 (=Col C)':<15}")
print("-" * 50)

old_row4 = {'AXWF6': 116, 'AXWG6': 'VALUEERROR', 'AXWH6': 44.5, 'AXWJ6': 48, 
            'AXWK6': 110, 'AXWM6': 99, 'AXWU6': 77, 'AXWZ6': 55.5,
            'AXWH7': 81.5, 'AXWM7': 66.5, 'AXWU7': 69, 'AXWZ7': 75}

for contract in column_contracts:
    old_val = old_row4.get(contract, '?')
    new_val = contracts[contract][1]
    changed = " ← CHANGED" if old_val != new_val else ""
    print(f"{contract:<10} {str(old_val):<12} {str(new_val):<15}{changed}")

print("\n" + "="*60)
print("Row 4 cells are highlighted GREEN to show corrected values")
print("="*60)
